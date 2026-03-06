from __future__ import annotations

import argparse
import html
import json
import math
import os
import re
import sys
import time
import traceback
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlparse

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    from zoneinfo import ZoneInfo  # py3.9+
except Exception:
    ZoneInfo = None  # type: ignore


# ============================================================
# Regex / constants
# ============================================================

TOKEN_RE = re.compile(r"/challenge/([A-Za-z0-9_-]+)")
ISO_RE = re.compile(r"^\d{4}-\d{2}-\d{2}")
EPOCH_RE = re.compile(r"^\d{10,13}$")
SETTING_LABEL_RE = re.compile(r'game-settings-list_settingLabel[^"]*">(.*?)</div>', re.S)

DEFAULT_TZ = "Europe/Stockholm"
DEFAULT_INFORMATION_CONFIG_NAME = "information_config.json"

# Fixed weekly map slots (index -> category)
MAP_SLOT_KEY_BY_INDEX = {
    1: "moving_1",
    2: "moving_2",
    3: "no_move_1",
    4: "no_move_2",
    5: "nmpz_1",
    6: "nmpz_2",
}

SLOT_KEYS_ORDER = ["moving_1", "moving_2", "no_move_1", "no_move_2", "nmpz_1", "nmpz_2"]

SLOT_LABEL_BY_KEY = {
    "moving_1": "Moving 1",
    "moving_2": "Moving 2",
    "no_move_1": "No move 1",
    "no_move_2": "No move 2",
    "nmpz_1": "NMPZ 1",
    "nmpz_2": "NMPZ 2",
}

SUBLEAGUE_SLOT_KEYS = {
    "Moving": ["moving_1", "moving_2"],
    "No move": ["no_move_1", "no_move_2"],
    "NMPZ": ["nmpz_1", "nmpz_2"],
    "Sverige": ["moving_1", "no_move_2"],
}

DEFAULT_INFORMATION_ROWS = [
    "Ingen anmälan krävs - det är bara att spela veckans challenges!",
    "För att öppna länken: klicka på den understrukna raden i varje kolumn. Exempelvis \"🔗 Moving 1 | Moving - 3 min\".",
    "Preliminära poäng utdelas under veckan. De kan gå upp beroende på hur många spelare som placerar sig under dig.",
    "Poäng delas ut enligt pro league-systemet: sista plats får 1 poäng, näst sista 2 poäng, tredje sista 3 poäng osv.",
    "Tiebreaker vid samma poäng är tid. Om två spelare delar plats får båda poäng för den delade placeringen.",
    "Varje vecka avslutas onsdag kl 20.00. Om poängen inte är ihopräknade då kan du spela tills poängen är ihopräknade.",
    "Vid frågor, skriv i #ligan.",
]

# Excel styling
DARK = PatternFill("solid", fgColor="2B2B2B")
MID = PatternFill("solid", fgColor="3A3A3A")
ROW_A = PatternFill("solid", fgColor="D9EAD3")
ROW_B = PatternFill("solid", fgColor="C9E2BC")
WHITE = PatternFill("solid", fgColor="FFFFFF")
MEDAL_GOLD = PatternFill("solid", fgColor="FFD966")
MEDAL_SILVER = PatternFill("solid", fgColor="D9D9D9")
MEDAL_BRONZE = PatternFill("solid", fgColor="F4B183")

FONT_HDR = Font(color="FFFFFF", bold=True)
FONT_HDR_BIG = Font(color="FFFFFF", bold=True, size=16)
FONT_HDR_MED = Font(color="FFFFFF", bold=True, size=12)
FONT_BODY = Font(color="000000", bold=False)

THIN = Side(style="thin", color="1F1F1F")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ============================================================
# Data classes
# ============================================================

@dataclass
class WeekSpec:
    label: str
    urls_path: Path
    deadline: Optional[str] = None  # user string, parsed later


@dataclass
class Entry:
    week_label: str
    map_index: int
    map_url: str
    map_token: str
    map_name: str
    rule_text: str

    player: str
    total_pts: int
    total_time: int  # tie-breaker within map
    played_at_epoch: Optional[int]  # optional, for deadline filtering


# ============================================================
# CLI
# ============================================================

def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    ap = argparse.ArgumentParser()

    # repeatable week spec:
    #   --week "Vecka 1|urls_week1.txt|2026-02-18 20:00"
    # deadline is optional
    ap.add_argument(
        "--week",
        action="append",
        default=[],
        help='Repeatable. Format: "LABEL|URLS_FILE|DEADLINE". Deadline optional. Example: --week "Vecka 2|urls_week2.txt|2026-02-25 20:00"',
    )

    ap.add_argument("--out-base", default="Liga_overview", help="Base filename without extension.")
    ap.add_argument(
        "--information-config",
        default="",
        help=f"Path to JSON config for Information sheet. Default behavior uses ./{DEFAULT_INFORMATION_CONFIG_NAME} when present.",
    )
    ap.add_argument("--tz", default=DEFAULT_TZ, help="Timezone for deadlines, e.g. Europe/Stockholm")
    ap.add_argument("--ncfa", default="", help="Override GEOGUESSR_NCFA env var")
    ap.add_argument("--timeout", type=float, default=30.0)

    # scoring / tie handling (only for exact ties on points+time; rare)
    ap.add_argument("--tie", default="average", choices=["average", "dense", "min", "max"])

    # highscores pagination
    ap.add_argument("--page-size", type=int, default=200)
    ap.add_argument("--max-players", type=int, default=5000)

    # played-at filtering
    ap.add_argument("--fetch-played-at", action="store_true", help="Try to fetch played timestamp per entry via extra API calls.")
    ap.add_argument("--keep-missing-time", action="store_true", help="When filtering, keep entries where played_at cannot be determined (default: exclude).")

    # debug
    ap.add_argument("--debug", action="store_true")
    ap.add_argument("--dump-json", action="store_true", help="Dump first highscores payload per map into ./debug_json/")

    return ap.parse_args(argv)


# ============================================================
# Core helpers
# ============================================================

def debug_print(debug: bool, *args):
    if debug:
        print(*args)


def extract_token(url: str) -> str:
    m = TOKEN_RE.search(url)
    if m:
        return m.group(1)
    p = urlparse(url).path.rstrip("/").split("/")
    if not p or not p[-1]:
        raise ValueError(f"Could not extract token from URL: {url}")
    return p[-1]


def load_urls(path: Path) -> List[str]:
    txt = path.read_text(encoding="utf-8")
    out: List[str] = []
    for line in txt.splitlines():
        s = line.strip()
        if s and not s.startswith("#"):
            out.append(s)
    return out


def make_session(ncfa: str) -> requests.Session:
    s = requests.Session()
    s.headers.update(
        {
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/json, text/plain, */*",
            "Referer": "https://www.geoguessr.com/",
            "Origin": "https://www.geoguessr.com",
        }
    )
    s.cookies.set("_ncfa", ncfa, domain=".geoguessr.com", path="/")
    return s


def http_get_json(session: requests.Session, url: str, timeout: float, debug: bool) -> Any:
    r = session.get(url, timeout=timeout)
    debug_print(debug, f"[HTTP] GET {url} -> {r.status_code} len={len(r.text)}")
    if r.status_code >= 400:
        snippet = r.text[:300].replace("\n", "\\n")
        raise RuntimeError(f"HTTP {r.status_code} for {url}: {snippet}")
    return r.json()


def http_get_text(session: requests.Session, url: str, timeout: float, debug: bool) -> str:
    r = session.get(url, timeout=timeout)
    debug_print(debug, f"[HTTP] GET {url} -> {r.status_code} len={len(r.text)}")
    if r.status_code >= 400:
        snippet = r.text[:300].replace("\n", "\\n")
        raise RuntimeError(f"HTTP {r.status_code} for {url}: {snippet}")
    return r.text


def _parse_int_maybe(x: Any) -> Optional[int]:
    if x is None or isinstance(x, bool):
        return None
    if isinstance(x, int):
        return x
    if isinstance(x, float):
        return int(x)
    if isinstance(x, str):
        s = x.strip().replace(",", "")
        if s.isdigit():
            return int(s)
    return None


def _parse_time_limit_seconds(label: str) -> Optional[int]:
    txt = label.strip().lower()
    if not txt:
        return None
    m = re.search(r"(\d+)\s*min", txt)
    if m:
        return int(m.group(1)) * 60
    m = re.search(r"(\d+)\s*(?:s|sec|secs|second|seconds)\b", txt)
    if m:
        return int(m.group(1))
    return None


def _clean_setting_label(raw: str) -> str:
    txt = re.sub(r"<!--.*?-->", "", raw, flags=re.S)
    txt = re.sub(r"<[^>]+>", "", txt)
    txt = html.unescape(txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def map_slot_key(map_index: Any) -> str:
    idx = _parse_int_maybe(map_index)
    if idx is None:
        return "unknown"
    return MAP_SLOT_KEY_BY_INDEX.get(idx, f"map_{idx}")


def map_slot_label(slot_key: str) -> str:
    return SLOT_LABEL_BY_KEY.get(slot_key, slot_key.replace("_", " ").title())


def default_information_rows() -> List[str]:
    return list(DEFAULT_INFORMATION_ROWS)


def _normalize_information_rows(rows: Any) -> List[str]:
    if not isinstance(rows, list):
        return default_information_rows()
    out: List[str] = []
    for row in rows:
        if not isinstance(row, str):
            continue
        txt = row.strip()
        if txt:
            out.append(txt)
    return out or default_information_rows()


def load_information_rows(config_path: Optional[Path], debug: bool = False) -> List[str]:
    if config_path is None or not config_path.exists():
        return default_information_rows()
    try:
        payload = json.loads(config_path.read_text(encoding="utf-8"))
    except Exception as e:
        debug_print(debug, f"[INFO-CONFIG] failed to parse {config_path}: {e}")
        return default_information_rows()

    if isinstance(payload, dict):
        return _normalize_information_rows(payload.get("information_rows"))
    if isinstance(payload, list):
        return _normalize_information_rows(payload)
    return default_information_rows()


# ============================================================
# Highscores parsing (schema based on your payload)
# ============================================================

def extract_items(payload: Any) -> List[dict]:
    if isinstance(payload, dict):
        v = payload.get("items")
        if isinstance(v, list) and (not v or isinstance(v[0], dict)):
            return v  # type: ignore
    if isinstance(payload, list) and (not payload or isinstance(payload[0], dict)):
        return payload  # type: ignore
    raise RuntimeError("Could not locate highscores items list.")


def rule_text_from_game(game: dict) -> str:
    forbid_moving = game.get("forbidMoving")
    forbid_zooming = game.get("forbidZooming")
    forbid_rotating = game.get("forbidRotating")
    time_limit = game.get("timeLimit")

    parts: List[str] = []

    if forbid_moving is True and forbid_zooming is True and forbid_rotating is True:
        parts.append("NMPZ")
    elif forbid_moving is True:
        # could be NM or NMP depending on other flags
        if forbid_rotating is True and forbid_zooming is False:
            parts.append("NMP")
        elif forbid_rotating is True and forbid_zooming is True:
            parts.append("NMPZ")
        else:
            parts.append("NM")
    else:
        parts.append("Moving")

    if isinstance(time_limit, int) and time_limit > 0:
        if time_limit % 60 == 0:
            parts.append(f"{time_limit//60} min")
        else:
            parts.append(f"{time_limit}s")

    return " - ".join(parts)


def player_name_from_item(item: dict) -> str:
    try:
        nick = item["game"]["player"].get("nick")
        if isinstance(nick, str) and nick.strip():
            return nick.strip()
    except Exception:
        pass
    return "UNKNOWN"


def total_points_from_item(item: dict) -> int:
    # prefer totalScore.amount (string)
    try:
        amt = item["game"]["player"]["totalScore"].get("amount")
        v = _parse_int_maybe(amt)
        if v is not None:
            return v
    except Exception:
        pass

    # fallback numeric variants
    try:
        v = _parse_int_maybe(item["game"]["player"].get("totalScoreInPoints"))
        if v is not None:
            return v
    except Exception:
        pass

    return 0


def total_time_from_item(item: dict) -> int:
    # tie-break: lower is better
    try:
        v = _parse_int_maybe(item["game"]["player"].get("totalTime"))
        if v is not None:
            return v
    except Exception:
        pass
    return 10**12


def map_name_from_item(item: dict) -> str:
    try:
        name = item["game"].get("mapName")
        if isinstance(name, str) and name.strip():
            return name.strip()
    except Exception:
        pass
    return ""


# ============================================================
# Ranking + Borda with time tie-break
# ============================================================

def compute_rank_and_borda_with_time(
    pts_by_player: Dict[str, int],
    time_by_player: Dict[str, int],
    tie_mode: str,
) -> Tuple[Dict[str, float], Dict[str, float]]:
    """
    Ranking:
      - higher points is better
      - if equal points: lower totalTime is better
      - if equal points AND time: tie_mode decides rank for exact ties

    Returns:
      rank_best: 1.0 = best
      borda: N = best ... 1 = worst (fractional if average tie)
    """
    if not pts_by_player:
        return {}, {}

    players = list(pts_by_player.keys())
    players_sorted = sorted(players, key=lambda p: (-pts_by_player[p], time_by_player.get(p, 10**12)))

    groups: Dict[Tuple[int, int], List[str]] = {}
    for p in players_sorted:
        key = (pts_by_player[p], time_by_player.get(p, 10**12))
        groups.setdefault(key, []).append(p)

    keys_sorted = sorted(groups.keys(), key=lambda k: (-k[0], k[1]))

    rank_best: Dict[str, float] = {}
    current_rank = 1

    for key in keys_sorted:
        names = groups[key]
        k = len(names)
        occupied = list(range(current_rank, current_rank + k))  # +k (NOT +k+1)

        if k == 1:
            rank_best[names[0]] = float(current_rank)
            current_rank += 1
            continue

        if tie_mode == "average":
            val = sum(occupied) / len(occupied)
        elif tie_mode == "dense":
            val = float(current_rank)
        elif tie_mode == "min":
            val = float(min(occupied))
        elif tie_mode == "max":
            val = float(max(occupied))
        else:
            raise ValueError(tie_mode)

        for n in names:
            rank_best[n] = float(val)

        current_rank = current_rank + (1 if tie_mode == "dense" else k)

    N = len(pts_by_player)
    borda = {p: float(N - rank_best[p] + 1) for p in rank_best}
    return rank_best, borda


# ============================================================
# Played-at extraction (best-effort)
# ============================================================

def _iter_all_dicts(obj: Any) -> Iterable[dict]:
    if isinstance(obj, dict):
        yield obj
        for v in obj.values():
            yield from _iter_all_dicts(v)
    elif isinstance(obj, list):
        for it in obj:
            yield from _iter_all_dicts(it)


def _try_parse_epoch(val: Any) -> Optional[int]:
    # epoch seconds or ms, or ISO string
    if val is None:
        return None

    if isinstance(val, (int, float)) and not isinstance(val, bool):
        x = int(val)
        # if ms
        if x > 10_000_000_000:
            return x // 1000
        # seconds
        if x > 1_000_000_000:
            return x
        return None

    if isinstance(val, str):
        s = val.strip()
        if EPOCH_RE.match(s):
            x = int(s)
            if x > 10_000_000_000:
                return x // 1000
            if x > 1_000_000_000:
                return x
        if ISO_RE.match(s):
            # very light ISO parsing without extra deps: handle "YYYY-MM-DDTHH:MM:SSZ" or with offset
            try:
                # Python can parse many ISO formats via fromisoformat, but "Z" needs replacement
                ss = s.replace("Z", "+00:00")
                dt = pd.to_datetime(ss, utc=True)
                if pd.isna(dt):
                    return None
                return int(dt.timestamp())
            except Exception:
                return None
    return None


def extract_played_at_epoch(game_payload: Any) -> Optional[int]:
    """
    Best-effort scan for typical timestamp keys:
      createdAt, created, updatedAt, finishedAt, endedAt, startTime, endTime, completedAt, etc.
    Prioritizes end/finish timestamps over generic updated/timestamp fields.
    """
    def key_priority(lk: str) -> Optional[int]:
        if lk in {"finishedat", "endedat", "endtime", "completedat", "completed", "finished", "ended"}:
            return 0
        if lk in {"createdat", "created", "startedat", "starttime", "started"}:
            return 1
        if any(x in lk for x in ["finished", "ended", "completed", "end"]):
            return 2
        if any(x in lk for x in ["created", "started", "start"]):
            return 3
        if lk in {"timestamp", "time"}:
            return 4
        if lk in {"updatedat", "updated"} or "updated" in lk:
            return 5
        return None

    best: Optional[Tuple[int, int]] = None  # (priority, epoch)

    for d in _iter_all_dicts(game_payload):
        for k, v in d.items():
            lk = str(k).lower()
            prio = key_priority(lk)
            if prio is None:
                continue

            ep = _try_parse_epoch(v)
            if ep is None:
                continue

            # Prefer lower priority class; within same class pick latest epoch.
            if best is None or prio < best[0] or (prio == best[0] and ep > best[1]):
                best = (prio, ep)

    return best[1] if best is not None else None


def fetch_game_details_for_played_at(
    session: requests.Session,
    game_token: str,
    timeout: float,
    debug: bool,
) -> Optional[int]:
    """
    Try a few endpoints. GeoGuessr may change schemas.
    We keep this tolerant: if an endpoint fails, try next.
    """
    endpoints = [
        f"https://www.geoguessr.com/api/v3/games/{game_token}",
        f"https://www.geoguessr.com/api/v3/results/{game_token}",
    ]
    for url in endpoints:
        try:
            payload = http_get_json(session, url, timeout=timeout, debug=debug)
            ep = extract_played_at_epoch(payload)
            if ep is not None:
                return ep
        except Exception as e:
            debug_print(debug, f"[played_at] endpoint failed: {url} -> {e}")
            continue
    return None


def fetch_challenge_landing_meta(
    session: requests.Session,
    challenge_token: str,
    timeout: float,
    debug: bool,
) -> Tuple[str, str]:
    """
    Best-effort parse from challenge landing page HTML.
    Useful when no highscores exist yet.
    Returns: (map_name, rule_text)
    """
    url = f"https://www.geoguessr.com/challenge/{challenge_token}"
    try:
        html_txt = http_get_text(session, url, timeout=timeout, debug=debug)
    except Exception as e:
        debug_print(debug, f"[landing-meta] failed for {challenge_token}: {e}")
        return "", ""

    labels_raw = SETTING_LABEL_RE.findall(html_txt)
    labels = [_clean_setting_label(x) for x in labels_raw if _clean_setting_label(x)]
    if not labels:
        return "", ""

    map_name = labels[0] if len(labels) >= 1 else ""
    time_label = labels[1] if len(labels) >= 2 else ""
    labels_lc = [x.lower() for x in labels]

    moving_allowed = any("moving allowed" in t for t in labels_lc)
    moving_not_allowed = any("moving not allowed" in t for t in labels_lc)
    panning_allowed = any("panning allowed" in t for t in labels_lc)
    panning_not_allowed = any("panning not allowed" in t for t in labels_lc)
    zooming_allowed = any("zooming allowed" in t for t in labels_lc)
    zooming_not_allowed = any("zooming not allowed" in t for t in labels_lc)

    mode = ""
    if moving_not_allowed:
        if panning_not_allowed and zooming_not_allowed:
            mode = "NMPZ"
        elif panning_not_allowed:
            mode = "NMP"
        else:
            mode = "NM"
    elif moving_allowed:
        mode = "Moving"

    secs = _parse_time_limit_seconds(time_label)
    time_part = ""
    if secs is not None:
        if secs % 60 == 0:
            time_part = f"{secs // 60} min"
        else:
            time_part = f"{secs}s"
    elif time_label:
        time_part = time_label

    parts = [p for p in [mode, time_part] if p]
    rule_text = " - ".join(parts)

    # If flags were present but mode wasn't resolved, infer from negation labels.
    if not rule_text and (panning_allowed or panning_not_allowed or zooming_allowed or zooming_not_allowed):
        forbid_moving = moving_not_allowed
        forbid_rotating = panning_not_allowed
        forbid_zooming = zooming_not_allowed
        pseudo_game = {
            "forbidMoving": forbid_moving,
            "forbidRotating": forbid_rotating,
            "forbidZooming": forbid_zooming,
            "timeLimit": secs,
        }
        rule_text = rule_text_from_game(pseudo_game)

    return map_name, rule_text


# ============================================================
# Highscores fetch
# ============================================================

def fetch_highscores_items(
    session: requests.Session,
    challenge_token: str,
    timeout: float,
    debug: bool,
    page_size: int,
    max_players: int,
) -> List[dict]:
    all_items: List[dict] = []
    offset = 0
    while True:
        url = (
            f"https://www.geoguessr.com/api/v3/results/highscores/{challenge_token}"
            f"?friends=false&limit={page_size}&offset={offset}"
        )
        payload = http_get_json(session, url, timeout=timeout, debug=debug)
        items = extract_items(payload)
        if not items:
            break
        all_items.extend(items)
        if len(items) < page_size:
            break
        offset += page_size
        if offset >= max_players:
            break
    return all_items


# ============================================================
# Deadline parsing
# ============================================================

def parse_deadline_epoch(deadline_str: str, tz_name: str) -> int:
    """
    Accepts e.g. "2026-02-25 20:00" or ISO.
    Interprets as tz_name local time.
    """
    if ZoneInfo is None:
        raise RuntimeError("zoneinfo not available. Use Python 3.9+ or install backports.zoneinfo.")

    try:
        tz = ZoneInfo(tz_name)
    except Exception as e:
        raise ValueError(
            f'Unknown timezone "{tz_name}". Example valid value: "Europe/Stockholm".'
        ) from e

    # Use pandas only for string parsing. Do timezone attachment/conversion via stdlib
    # to avoid pandas+zoneinfo incompatibilities in some environments.
    dt = pd.to_datetime(deadline_str)
    if pd.isna(dt):
        raise ValueError(f"Could not parse deadline: {deadline_str}")

    py_dt = dt.to_pydatetime() if hasattr(dt, "to_pydatetime") else dt
    if getattr(py_dt, "tzinfo", None) is None:
        py_dt = py_dt.replace(tzinfo=tz)
    else:
        py_dt = py_dt.astimezone(tz)
    # convert to epoch seconds
    return int(py_dt.timestamp())


# ============================================================
# Build entries per week
# ============================================================

def build_week_entries(
    session: requests.Session,
    week: WeekSpec,
    tz_name: str,
    timeout: float,
    debug: bool,
    dump_json: bool,
    page_size: int,
    max_players: int,
    fetch_played_at: bool,
) -> Tuple[List[Entry], List[dict], bool, int]:
    """
    Returns (entries, map_meta_rows, has_any_played_at, failed_maps_count).
    """
    urls = load_urls(week.urls_path)
    if not urls:
        raise RuntimeError(f"{week.urls_path} is empty")

    out_entries: List[Entry] = []
    map_meta_rows: List[dict] = []
    has_any_played_at = False
    failed_maps_count = 0

    debug_dir = week.urls_path.parent / "debug_json"
    if dump_json:
        debug_dir.mkdir(parents=True, exist_ok=True)

    played_at_cache: Dict[str, Optional[int]] = {}

    for map_idx, url in enumerate(urls, start=1):
        token = extract_token(url)
        map_name = f"Map {map_idx}"
        rule_text = ""
        try:
            items = fetch_highscores_items(
                session=session,
                challenge_token=token,
                timeout=timeout,
                debug=debug,
                page_size=page_size,
                max_players=max_players,
            )
        except Exception as e:
            failed_maps_count += 1
            print(f"[WARN] {week.label} map {map_idx}: kunde inte hämta resultat för {url} ({e})")
            landing_name, landing_rule = fetch_challenge_landing_meta(session, token, timeout=timeout, debug=debug)
            if landing_name:
                map_name = landing_name
            if landing_rule:
                rule_text = landing_rule
            map_meta_rows.append(
                {
                    "week": week.label,
                    "map_index": map_idx,
                    "map_url": url,
                    "map_name": map_name,
                    "rule_text": rule_text,
                }
            )
            continue

        if dump_json:
            p = debug_dir / f"{week.label.replace(' ', '_')}_map{map_idx}_highscores.json"
            p.write_text(json.dumps({"token": token, "items": items}, ensure_ascii=False, indent=2), encoding="utf-8")

        # map info from first item (stable in your payload)
        if items:
            try:
                game0 = items[0]["game"]
                map_name = str(game0.get("mapName") or "").strip()
                rule_text = rule_text_from_game(game0)
            except Exception:
                rule_text = ""
        else:
            landing_name, landing_rule = fetch_challenge_landing_meta(session, token, timeout=timeout, debug=debug)
            if landing_name:
                map_name = landing_name
            if landing_rule:
                rule_text = landing_rule

        map_meta_rows.append(
            {
                "week": week.label,
                "map_index": map_idx,
                "map_url": url,
                "map_name": map_name or f"Map {map_idx}",
                "rule_text": rule_text or "",
            }
        )

        for it in items:
            if not isinstance(it, dict) or "game" not in it:
                continue
            name = player_name_from_item(it)
            if name == "UNKNOWN":
                continue
            pts = total_points_from_item(it)
            ttime = total_time_from_item(it)

            # played_at: requires extra call using game token
            played_at: Optional[int] = None
            if fetch_played_at:
                try:
                    game_token = it["game"].get("token")
                except Exception:
                    game_token = None

                if isinstance(game_token, str) and game_token:
                    if game_token in played_at_cache:
                        played_at = played_at_cache[game_token]
                    else:
                        played_at = fetch_game_details_for_played_at(session, game_token, timeout=timeout, debug=debug)
                        played_at_cache[game_token] = played_at

            if played_at is not None:
                has_any_played_at = True

            out_entries.append(
                Entry(
                    week_label=week.label,
                    map_index=map_idx,
                    map_url=url,
                    map_token=token,
                    map_name=map_name or f"Map {map_idx}",
                    rule_text=rule_text or "",
                    player=name,
                    total_pts=pts,
                    total_time=ttime,
                    played_at_epoch=played_at,
                )
            )

    return out_entries, map_meta_rows, has_any_played_at, failed_maps_count


# ============================================================
# Filtering + scoring aggregation
# ============================================================

def filter_entries_by_deadlines(
    entries: List[Entry],
    deadlines_epoch_by_week: Dict[str, int],
    keep_missing_time: bool,
    now_epoch: Optional[int] = None,
) -> List[Entry]:
    if now_epoch is None:
        now_epoch = int(time.time())

    out: List[Entry] = []
    for e in entries:
        dl = deadlines_epoch_by_week.get(e.week_label)
        if dl is None:
            # no deadline specified for this week => keep
            out.append(e)
            continue

        # Ongoing/future week: never filter out yet.
        if dl > now_epoch:
            out.append(e)
            continue

        if e.played_at_epoch is None:
            if keep_missing_time:
                out.append(e)
            continue

        if e.played_at_epoch <= dl:
            out.append(e)
    return out


def compute_week_tables(entries: List[Entry], tie_mode: str, map_meta_rows: Optional[List[dict]] = None) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Returns:
      df_overview_rows: per entry with rank/borda within each (week,map)
      df_weekly: weekly summary per player (sum of borda)
      df_week_meta: map meta per (week,map) for headers
    """
    if not entries:
        if map_meta_rows:
            df_week_meta = pd.DataFrame(map_meta_rows, columns=["week", "map_index", "map_url", "map_name", "rule_text"])
            if not df_week_meta.empty:
                df_week_meta = (
                    df_week_meta.drop_duplicates(subset=["week", "map_index"], keep="last")
                    .sort_values(["week", "map_index"])
                    .reset_index(drop=True)
                )
        else:
            df_week_meta = pd.DataFrame(columns=["week", "map_index", "map_url", "map_name", "rule_text"])
        return (
            pd.DataFrame(columns=["week", "map_index", "map_url", "map_name", "rule_text", "slot_key", "slot_label", "player", "total_pts", "total_time", "rank_best", "borda_points", "played_at_epoch"]),
            pd.DataFrame(columns=["week", "player", "weekly_borda", "weekly_total_pts", "maps_counted"]),
            df_week_meta,
        )

    df = pd.DataFrame([{
        "week": e.week_label,
        "map_index": e.map_index,
        "map_url": e.map_url,
        "map_token": e.map_token,
        "map_name": e.map_name,
        "rule_text": e.rule_text,
        "player": e.player,
        "total_pts": e.total_pts,
        "total_time": e.total_time,
        "played_at_epoch": e.played_at_epoch,
    } for e in entries])
    df["slot_key"] = df["map_index"].apply(map_slot_key)
    df["slot_label"] = df["slot_key"].apply(map_slot_label)

    # discovered meta per map from result payloads
    df_week_meta_seen = (
        df[["week", "map_index", "map_url", "map_name", "rule_text"]]
        .drop_duplicates()
        .sort_values(["week", "map_index"])
        .reset_index(drop=True)
    )

    if map_meta_rows:
        df_week_meta_base = pd.DataFrame(map_meta_rows, columns=["week", "map_index", "map_url", "map_name", "rule_text"])
        if not df_week_meta_base.empty:
            df_week_meta_base = (
                df_week_meta_base.drop_duplicates(subset=["week", "map_index"], keep="last")
                .sort_values(["week", "map_index"])
                .reset_index(drop=True)
            )
    else:
        df_week_meta_base = pd.DataFrame(columns=["week", "map_index", "map_url", "map_name", "rule_text"])

    if df_week_meta_base.empty:
        df_week_meta = df_week_meta_seen
    else:
        df_week_meta = df_week_meta_base.merge(
            df_week_meta_seen.rename(
                columns={
                    "map_url": "seen_map_url",
                    "map_name": "seen_map_name",
                    "rule_text": "seen_rule_text",
                }
            ),
            on=["week", "map_index"],
            how="left",
        )

        def _prefer_non_empty(base: Any, seen: Any, fallback: str) -> str:
            b = str(base).strip() if isinstance(base, str) else ""
            s = str(seen).strip() if isinstance(seen, str) else ""
            if s:
                return s
            if b:
                return b
            return fallback

        df_week_meta["map_url"] = [
            _prefer_non_empty(b, s, "")
            for b, s in zip(df_week_meta.get("map_url", pd.Series(dtype=str)), df_week_meta.get("seen_map_url", pd.Series(dtype=str)))
        ]
        df_week_meta["map_name"] = [
            _prefer_non_empty(b, s, f"Map {int(mi)}")
            for b, s, mi in zip(
                df_week_meta.get("map_name", pd.Series(dtype=str)),
                df_week_meta.get("seen_map_name", pd.Series(dtype=str)),
                df_week_meta["map_index"],
            )
        ]
        df_week_meta["rule_text"] = [
            _prefer_non_empty(b, s, "")
            for b, s in zip(df_week_meta.get("rule_text", pd.Series(dtype=str)), df_week_meta.get("seen_rule_text", pd.Series(dtype=str)))
        ]
        df_week_meta = df_week_meta[["week", "map_index", "map_url", "map_name", "rule_text"]]

    # compute rank/borda within each week+map
    out_rows: List[dict] = []

    for (w, mi), g in df.groupby(["week", "map_index"], sort=True):
        pts_map = {row["player"]: int(row["total_pts"]) for _, row in g.iterrows()}
        time_map = {row["player"]: int(row["total_time"]) for _, row in g.iterrows()}
        rank_best, borda = compute_rank_and_borda_with_time(pts_map, time_map, tie_mode=tie_mode)

        for _, row in g.iterrows():
            p = row["player"]
            out_rows.append({
                **row.to_dict(),
                "rank_best": rank_best.get(p),
                "borda_points": borda.get(p),
            })

    df_overview = pd.DataFrame(out_rows)

    # weekly summary: sum borda across maps (and keep raw points sum too)
    df_weekly = (
        df_overview.groupby(["week", "player"], as_index=False)
        .agg(
            weekly_borda=("borda_points", "sum"),
            weekly_total_pts=("total_pts", "sum"),
            maps_counted=("map_index", "nunique"),
        )
        .sort_values(["week", "weekly_borda", "weekly_total_pts"], ascending=[True, False, False])
        .reset_index(drop=True)
    )

    return df_overview, df_weekly, df_week_meta


def compute_total_tables(df_overview: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Total standings + stats.
    """
    base_cols_total = [
        "player",
        "total_borda",
        "total_pts",
        "maps_counted",
        "weeks_counted",
        "avg_borda_per_map",
        "avg_borda_per_week",
        "avg_pts_per_map",
        "cat_moving_1",
        "cat_moving_2",
        "cat_no_move_1",
        "cat_no_move_2",
        "cat_nmpz_1",
        "cat_nmpz_2",
        "cat_moving",
        "cat_no_move",
        "cat_nmpz",
        "cat_sverige",
    ]
    base_cols_stats = base_cols_total + ["best_week", "best_week_borda", "best_week_pts"]

    if df_overview.empty:
        total = pd.DataFrame(columns=base_cols_total)
        stats = pd.DataFrame(columns=base_cols_stats)
        return total, stats

    dfo = df_overview.copy()
    dfo["week_map_key"] = dfo["week"].astype(str) + "::" + dfo["map_index"].astype(str)
    dfo["slot_key"] = dfo["map_index"].apply(map_slot_key)

    by_player = (
        dfo.groupby("player", as_index=False)
        .agg(
            total_borda=("borda_points", "sum"),
            total_pts=("total_pts", "sum"),
            maps_counted=("week_map_key", "nunique"),
            weeks_counted=("week", "nunique"),
        )
    )
    by_player["avg_borda_per_map"] = by_player["total_borda"] / by_player["maps_counted"].clip(lower=1)
    by_player["avg_borda_per_week"] = by_player["total_borda"] / by_player["weeks_counted"].clip(lower=1)
    by_player["avg_pts_per_map"] = by_player["total_pts"] / by_player["maps_counted"].clip(lower=1)

    # Slot totals
    slot_scores = (
        dfo[dfo["slot_key"].isin(SLOT_KEYS_ORDER)]
        .groupby(["player", "slot_key"], as_index=False)
        .agg(slot_borda=("borda_points", "sum"))
    )
    if not slot_scores.empty:
        slot_pivot = (
            slot_scores.pivot_table(index="player", columns="slot_key", values="slot_borda", aggfunc="sum")
            .fillna(0.0)
            .reset_index()
        )
        by_player = by_player.merge(slot_pivot, on="player", how="left")

    for key in SLOT_KEYS_ORDER:
        if key not in by_player.columns:
            by_player[key] = 0.0

    by_player = by_player.rename(columns={
        "moving_1": "cat_moving_1",
        "moving_2": "cat_moving_2",
        "no_move_1": "cat_no_move_1",
        "no_move_2": "cat_no_move_2",
        "nmpz_1": "cat_nmpz_1",
        "nmpz_2": "cat_nmpz_2",
    })

    by_player["cat_moving"] = by_player["cat_moving_1"] + by_player["cat_moving_2"]
    by_player["cat_no_move"] = by_player["cat_no_move_1"] + by_player["cat_no_move_2"]
    by_player["cat_nmpz"] = by_player["cat_nmpz_1"] + by_player["cat_nmpz_2"]
    by_player["cat_sverige"] = by_player["cat_moving_1"] + by_player["cat_no_move_2"]

    total = by_player.sort_values(["total_borda", "total_pts"], ascending=[False, False]).reset_index(drop=True)
    total = total.reindex(columns=base_cols_total)

    # extra stats: best week, avg per week, etc.
    per_week = (
        dfo.groupby(["player", "week"], as_index=False)
        .agg(
            week_borda=("borda_points", "sum"),
            week_pts=("total_pts", "sum"),
            week_maps=("week_map_key", "nunique"),
        )
    )
    best_week = per_week.sort_values(["player", "week_borda", "week_pts"], ascending=[True, False, False]).groupby("player").head(1)
    best_week = best_week[["player", "week", "week_borda", "week_pts"]].rename(columns={"week": "best_week", "week_borda": "best_week_borda", "week_pts": "best_week_pts"})

    stats = total.merge(best_week, on="player", how="left")
    stats = stats.sort_values(["total_borda", "total_pts"], ascending=[False, False]).reset_index(drop=True)
    stats = stats.reindex(columns=base_cols_stats)

    return total, stats


def compute_subleague_tables(df_overview: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    out: Dict[str, pd.DataFrame] = {}
    if df_overview.empty:
        for league_name in SUBLEAGUE_SLOT_KEYS:
            out[league_name] = pd.DataFrame(columns=["player", "league_points", "total_pts", "maps_counted", "weeks_counted"])
        return out

    dfo = df_overview.copy()
    dfo["slot_key"] = dfo["map_index"].apply(map_slot_key)
    dfo["week_map_key"] = dfo["week"].astype(str) + "::" + dfo["map_index"].astype(str)

    for league_name, slots in SUBLEAGUE_SLOT_KEYS.items():
        part = dfo[dfo["slot_key"].isin(slots)]
        if part.empty:
            out[league_name] = pd.DataFrame(columns=["player", "league_points", "total_pts", "maps_counted", "weeks_counted"])
            continue

        table = (
            part.groupby("player", as_index=False)
            .agg(
                league_points=("borda_points", "sum"),
                total_pts=("total_pts", "sum"),
                maps_counted=("week_map_key", "nunique"),
                weeks_counted=("week", "nunique"),
            )
            .sort_values(["league_points", "total_pts"], ascending=[False, False])
            .reset_index(drop=True)
        )
        out[league_name] = table
    return out


# ============================================================
# Excel output
# ============================================================

def set_col_widths(ws, widths: Dict[int, float]) -> None:
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def style_cell(ws, r: int, c: int, *, fill=None, font=None, align=None, border=True) -> None:
    cell = ws.cell(r, c)
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border:
        cell.border = BORDER_THIN


def merge_and_style(ws, r1: int, c1: int, r2: int, c2: int, value: str, *, fill, font, align) -> None:
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1)
    cell.value = value
    cell.fill = fill
    cell.font = font
    cell.alignment = align
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = BORDER_THIN
            ws.cell(r, c).fill = fill


def rank_row_fill(rank: int, fallback_fill: PatternFill) -> PatternFill:
    if rank == 1:
        return MEDAL_GOLD
    if rank == 2:
        return MEDAL_SILVER
    if rank == 3:
        return MEDAL_BRONZE
    return fallback_fill


def write_week_sheet(
    wb: Workbook,
    week_label: str,
    deadline_str: str,
    df_week: pd.DataFrame,
    df_overview: pd.DataFrame,
    df_meta: pd.DataFrame,
) -> None:
    """
    Layout similar to your Google-sheet:
      Row 1: Week label / Map 1..N
      Row 2: Deadline / Map names
      Row 3: headers (#, Spelare, Poäng, links)
      Rows : ranking table with per-map borda points
    """
    ws = wb.create_sheet(week_label)

    # Determine map columns
    meta = df_meta[df_meta["week"] == week_label].sort_values("map_index")
    maps = meta.to_dict("records")
    n_maps = len(maps)

    col_rank = 1
    col_player = 2
    col_total = 3
    col_map_start = 4

    r1, r2, r3, r_data_start = 1, 2, 3, 4

    merge_and_style(ws, r1, 1, r1, 3, week_label, fill=DARK, font=FONT_HDR_BIG, align=CENTER)
    merge_and_style(ws, r2, 1, r2, 3, deadline_str, fill=DARK, font=FONT_HDR_MED, align=CENTER)

    for i in range(n_maps):
        c = col_map_start + i
        merge_and_style(ws, r1, c, r1, c, f"Map {i+1}", fill=DARK, font=FONT_HDR_MED, align=CENTER)
        merge_and_style(ws, r2, c, r2, c, str(maps[i].get("map_name") or f"Map {i+1}"), fill=DARK, font=FONT_HDR_MED, align=CENTER)

    ws.cell(r3, col_rank).value = "#"
    ws.cell(r3, col_player).value = "Spelare"
    ws.cell(r3, col_total).value = "Poäng"
    for c in (col_rank, col_player, col_total):
        style_cell(ws, r3, c, fill=MID, font=FONT_HDR, align=CENTER)

    # header links per map
    for i in range(n_maps):
        c = col_map_start + i
        slot_key = map_slot_key(maps[i].get("map_index"))
        slot_label = map_slot_label(slot_key)
        rule = str(maps[i].get("rule_text") or "")
        url = str(maps[i].get("map_url") or "")
        txt = f"🔗 {slot_label}"
        if rule:
            txt = f"{txt} | {rule}"

        cell = ws.cell(r3, c)
        cell.value = txt
        if url:
            cell.hyperlink = url
        cell.fill = MID
        cell.alignment = CENTER
        cell.border = BORDER_THIN
        cell.font = Font(color="FFFFFF", bold=True, underline="single")

    ws.freeze_panes = ws["A4"]
    ws.row_dimensions[r2].height = 38
    ws.row_dimensions[r3].height = 34

    # column widths
    widths = {
        col_rank: 4.5,
        col_player: 22.0,
        col_total: 8.0,
    }
    for i in range(n_maps):
        widths[col_map_start + i] = 20.0
    set_col_widths(ws, widths)

    # per-map pivot for this week
    dwo = df_overview[df_overview["week"] == week_label]
    pivot = pd.DataFrame()
    if not dwo.empty:
        pivot = dwo.pivot_table(index="player", columns="map_index", values="borda_points", aggfunc="max")

    # weekly order
    dw = df_week[df_week["week"] == week_label].sort_values(["weekly_borda", "weekly_total_pts"], ascending=[False, False])
    ordered = dw["player"].tolist()

    for idx, player in enumerate(ordered, start=1):
        r = r_data_start + (idx - 1)
        base_fill = ROW_A if (idx % 2 == 1) else ROW_B
        fill = rank_row_fill(idx, base_fill)

        ws.cell(r, col_rank).value = idx
        ws.cell(r, col_player).value = player
        pts = float(dw.loc[dw["player"] == player, "weekly_borda"].iloc[0])
        ws.cell(r, col_total).value = int(pts) if abs(pts - int(pts)) < 1e-9 else pts

        style_cell(ws, r, col_rank, fill=fill, font=FONT_BODY, align=CENTER)
        style_cell(ws, r, col_player, fill=fill, font=FONT_BODY, align=LEFT)
        style_cell(ws, r, col_total, fill=fill, font=Font(color="000000", bold=True), align=CENTER)

        for i in range(n_maps):
            c = col_map_start + i
            map_idx = int(_parse_int_maybe(maps[i].get("map_index")) or (i + 1))
            val = None
            if not pivot.empty and player in pivot.index and map_idx in pivot.columns:
                v = pivot.loc[player, map_idx]
                if pd.notna(v):
                    val = float(v)
            ws.cell(r, c).value = "" if val is None else (int(val) if abs(val - int(val)) < 1e-9 else val)
            style_cell(ws, r, c, fill=fill, font=FONT_BODY, align=CENTER)


def write_total_sheet(wb: Workbook, df_total: pd.DataFrame, df_overview: pd.DataFrame, weeks: List[str]) -> None:
    ws = wb.create_sheet("Total")

    # Header
    merge_and_style(ws, 1, 1, 1, 7 + len(weeks), "Totalställning", fill=DARK, font=FONT_HDR_BIG, align=CENTER)

    headers = ["#", "Spelare", "Poäng", "Total pts", "Snitt pts/karta", "Kartor", "Veckor"] + [f"{w}" for w in weeks]
    for c, h in enumerate(headers, start=1):
        ws.cell(2, c).value = h
        style_cell(ws, 2, c, fill=MID, font=FONT_HDR, align=CENTER)

    ws.freeze_panes = "A3"

    widths = {1: 4.5, 2: 22.0, 3: 14.0, 4: 10.0, 5: 14.0, 6: 8.0, 7: 8.0}
    for i in range(len(weeks)):
        widths[8 + i] = 12.0
    set_col_widths(ws, widths)

    # per-week totals pivot (borda)
    per_week = (
        df_overview.groupby(["player", "week"], as_index=False)
        .agg(week_borda=("borda_points", "sum"))
    )
    pivot = per_week.pivot_table(index="player", columns="week", values="week_borda", aggfunc="sum")

    for idx, row in enumerate(df_total.itertuples(index=False), start=1):
        r = 2 + idx
        base_fill = ROW_A if (idx % 2 == 1) else ROW_B
        fill = rank_row_fill(idx, base_fill)

        ws.cell(r, 1).value = idx
        ws.cell(r, 2).value = row.player
        ws.cell(r, 3).value = float(row.total_borda)
        ws.cell(r, 4).value = int(row.total_pts)
        ws.cell(r, 5).value = float(getattr(row, "avg_pts_per_map", 0) or 0)
        ws.cell(r, 6).value = int(row.maps_counted)
        ws.cell(r, 7).value = int(row.weeks_counted)

        for c in range(1, 8):
            style_cell(ws, r, c, fill=fill, font=FONT_BODY if c != 3 else Font(color="000000", bold=True), align=CENTER if c != 2 else LEFT)

        # week columns
        for j, w in enumerate(weeks):
            c = 8 + j
            val = ""
            if not pivot.empty and row.player in pivot.index and w in pivot.columns:
                v = pivot.loc[row.player, w]
                if pd.notna(v):
                    val = float(v)
                    val = int(val) if abs(val - int(val)) < 1e-9 else val
            ws.cell(r, c).value = val
            style_cell(ws, r, c, fill=fill, font=FONT_BODY, align=CENTER)


def write_stats_sheet(wb: Workbook, df_stats: pd.DataFrame) -> None:
    ws = wb.create_sheet("Stats")

    merge_and_style(ws, 1, 1, 1, 22, "Statistik", fill=DARK, font=FONT_HDR_BIG, align=CENTER)

    cols = [
        "#", "Spelare",
        "Poäng", "Total pts",
        "Kartor", "Veckor",
        "Moving 1", "Moving 2",
        "No move 1", "No move 2",
        "NMPZ 1", "NMPZ 2",
        "Moving", "No move", "NMPZ", "Sverige",
        "Snitt poäng / karta", "Snitt poäng / vecka", "Snitt pts / karta",
        "Bästa vecka", "Bästa vecka poäng", "Bästa vecka pts",
    ]

    for c, h in enumerate(cols, start=1):
        ws.cell(2, c).value = h
        style_cell(ws, 2, c, fill=MID, font=FONT_HDR, align=CENTER)

    ws.freeze_panes = "A3"

    widths = {
        1: 4.5, 2: 22.0,
        3: 12.0, 4: 10.0,
        5: 8.0, 6: 8.0,
        7: 10.0, 8: 10.0,
        9: 10.0, 10: 10.0,
        11: 10.0, 12: 10.0,
        13: 10.0, 14: 10.0, 15: 10.0, 16: 10.0,
        17: 14.0, 18: 14.0, 19: 12.0,
        20: 14.0, 21: 18.0, 22: 14.0,
    }
    set_col_widths(ws, widths)

    for idx, row in enumerate(df_stats.itertuples(index=False), start=1):
        r = 2 + idx
        base_fill = ROW_A if (idx % 2 == 1) else ROW_B
        fill = rank_row_fill(idx, base_fill)

        ws.cell(r, 1).value = idx
        ws.cell(r, 2).value = row.player
        ws.cell(r, 3).value = float(row.total_borda)
        ws.cell(r, 4).value = int(row.total_pts)
        ws.cell(r, 5).value = int(row.maps_counted)
        ws.cell(r, 6).value = int(row.weeks_counted)
        ws.cell(r, 7).value = float(getattr(row, "cat_moving_1", 0) or 0)
        ws.cell(r, 8).value = float(getattr(row, "cat_moving_2", 0) or 0)
        ws.cell(r, 9).value = float(getattr(row, "cat_no_move_1", 0) or 0)
        ws.cell(r, 10).value = float(getattr(row, "cat_no_move_2", 0) or 0)
        ws.cell(r, 11).value = float(getattr(row, "cat_nmpz_1", 0) or 0)
        ws.cell(r, 12).value = float(getattr(row, "cat_nmpz_2", 0) or 0)
        ws.cell(r, 13).value = float(getattr(row, "cat_moving", 0) or 0)
        ws.cell(r, 14).value = float(getattr(row, "cat_no_move", 0) or 0)
        ws.cell(r, 15).value = float(getattr(row, "cat_nmpz", 0) or 0)
        ws.cell(r, 16).value = float(getattr(row, "cat_sverige", 0) or 0)
        ws.cell(r, 17).value = float(row.avg_borda_per_map)
        ws.cell(r, 18).value = float(row.avg_borda_per_week)
        ws.cell(r, 19).value = float(row.avg_pts_per_map)
        ws.cell(r, 20).value = getattr(row, "best_week", "")
        ws.cell(r, 21).value = float(getattr(row, "best_week_borda", 0) or 0)
        ws.cell(r, 22).value = float(getattr(row, "best_week_pts", 0) or 0)

        for c in range(1, 23):
            align = LEFT if c == 2 else CENTER
            font = Font(color="000000", bold=True) if c in (3,) else FONT_BODY
            style_cell(ws, r, c, fill=fill, font=font, align=align)


def write_visualizations_sheet(
    wb: Workbook,
    df_overview: pd.DataFrame,
    df_total: pd.DataFrame,
    weeks: List[str],
    image_dir: Optional[Path] = None,
) -> None:
    ws = wb.create_sheet("Visualiseringar")
    merge_and_style(ws, 1, 1, 1, 24, "Visualiseringar (Experiment)", fill=DARK, font=FONT_HDR_BIG, align=CENTER)

    viz_names = [
        "V1A: Tid vs poäng (Moving)",
        "V1B: Tid vs poäng (No move)",
        "V1C: Tid vs poäng (NMPZ)",
        "V2: Total råpoäng topp 20",
        "V3: Snitt GeoGuessr-poäng per karttyp och spelare",
        "V4: Aktiva spelare per vecka",
        "V5: Ligans snitt GeoGuessr-poäng per underliga-kategori",
        "V6: Topp-spelare per karttyp (snitt GeoGuessr-poäng)",
        "V7: Poängfördelning per karttyp",
        "V8: Stabilitet vs nivå (std mot snittpoäng)",
        "V9: Veckotrend-heatmap (över/under eget snitt)",
        "V10: PCA spelarmönster + bidrag till PC1/PC2",
        "V11: Standardavvikelse per karttyp och spelare",
    ]
    ws["A3"] = "Diagramöversikt:"
    ws["A3"].font = Font(bold=True, color="1B314B")
    for i, txt in enumerate(viz_names, start=4):
        ws[f"A{i}"] = txt
        ws[f"A{i}"].font = Font(color="1B314B")

    set_col_widths(ws, {1: 58.0, 2: 14.0, 13: 4.0, 14: 58.0})

    if df_overview.empty or df_total.empty:
        ws["A16"] = "Ingen data tillgänglig för visualiseringar."
        ws["A16"].font = Font(color="AA0000", bold=True)
        return

    if image_dir is None:
        image_dir = Path("visualizations") / "latest"
    image_dir.mkdir(parents=True, exist_ok=True)
    for old_png in image_dir.glob("V*.png"):
        try:
            old_png.unlink()
        except Exception:
            pass

    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np
    except Exception as ex:
        ws["A16"] = f"Kunde inte skapa visualiseringsbilder (matplotlib saknas): {ex}"
        ws["A16"].font = Font(color="AA0000", bold=True)
        return

    def _empty_plot(ax, text: str = "Ingen data") -> None:
        ax.text(0.5, 0.5, text, ha="center", va="center", transform=ax.transAxes, fontsize=12)
        ax.set_xticks([])
        ax.set_yticks([])

    def _save_fig(fig, filename: str) -> Path:
        fig.tight_layout()
        out_path = image_dir / filename
        fig.savefig(out_path, dpi=150, bbox_inches="tight")
        plt.close(fig)
        return out_path

    def _insert_image(path: Path, anchor: str, width: int = 620, height: int = 320) -> None:
        img = XLImage(str(path))
        img.width = width
        img.height = height
        ws.add_image(img, anchor)

    def _detect_outlier_players(points_df: pd.DataFrame, x_col: str, y_col: str, z_thresh: float = 2.35) -> set[str]:
        if points_df.empty or len(points_df) < 5:
            return set()
        x = pd.to_numeric(points_df[x_col], errors="coerce").fillna(0.0).to_numpy(dtype=float)
        y = pd.to_numeric(points_df[y_col], errors="coerce").fillna(0.0).to_numpy(dtype=float)
        sx = float(np.std(x))
        sy = float(np.std(y))
        zx = np.zeros_like(x) if sx <= 1e-9 else (x - float(np.mean(x))) / sx
        zy = np.zeros_like(y) if sy <= 1e-9 else (y - float(np.mean(y))) / sy
        radial = np.sqrt(zx * zx + zy * zy)
        mask = radial >= z_thresh
        if "player" not in points_df.columns:
            return set()
        return set(points_df.loc[mask, "player"].astype(str).tolist())

    def _annotate_all_points(ax, x_vals: List[float], y_vals: List[float], labels: List[str], fontsize: int = 6) -> None:
        offsets = [(0, 0), (4, 2), (-4, 2), (4, -2), (-4, -2), (0, 4), (0, -4)]
        for i, (xv, yv, label) in enumerate(zip(x_vals, y_vals, labels)):
            dx, dy = offsets[i % len(offsets)]
            ax.annotate(
                str(label),
                (float(xv), float(yv)),
                textcoords="offset points",
                xytext=(dx, dy),
                fontsize=fontsize,
                color="#1F2D3D",
                bbox={"boxstyle": "round,pad=0.08", "fc": "white", "ec": "none", "alpha": 0.38},
            )

    dfo = df_overview.copy()
    dfo["slot_key"] = dfo["map_index"].apply(map_slot_key)
    dfo["slot_label"] = dfo["slot_key"].apply(map_slot_label)
    dfo["total_pts"] = pd.to_numeric(dfo["total_pts"], errors="coerce").fillna(0.0)
    dfo["total_time"] = pd.to_numeric(dfo["total_time"], errors="coerce").fillna(0.0)
    dfo["week_map_key"] = dfo["week"].astype(str) + "::" + dfo["map_index"].astype(str)

    slot_to_mode3 = {
        "moving_1": "Moving",
        "moving_2": "Moving",
        "no_move_1": "No move",
        "no_move_2": "No move",
        "nmpz_1": "NMPZ",
        "nmpz_2": "NMPZ",
    }
    dfo["mode3"] = dfo["slot_key"].map(slot_to_mode3).fillna("Other")

    weeks_seen = dfo["week"].dropna().astype(str).tolist()
    weeks_order = list(dict.fromkeys(list(weeks) + weeks_seen))

    total_maps = max(1, int(dfo["week_map_key"].nunique()))
    min_maps_for_labels = max(1, int(math.ceil(total_maps * 0.75)))
    maps_by_player = dfo.groupby("player")["week_map_key"].nunique()
    qualified_players = set(
        maps_by_player[maps_by_player >= min_maps_for_labels].index.astype(str).tolist()
    )

    week_participation = (
        dfo.groupby(["player", "week"], as_index=False)
        .agg(week_maps=("map_index", "nunique"))
    )
    expanded_rule = (
        week_participation.groupby("player", as_index=False)
        .agg(
            weeks_played=("week", "nunique"),
            max_week_maps=("week_maps", "max"),
            total_maps_played=("week_maps", "sum"),
        )
    )
    expanded_players = set(
        expanded_rule[
            (expanded_rule["max_week_maps"] >= 6)
            & (expanded_rule["weeks_played"] >= 2)
            & (expanded_rule["total_maps_played"] >= 7)
        ]["player"].astype(str).tolist()
    )

    total_pts_series = (
        df_total.set_index("player")["total_pts"]
        if ("player" in df_total.columns and "total_pts" in df_total.columns)
        else pd.Series(dtype=float)
    )
    total_pts_dict = {str(k): float(v) for k, v in total_pts_series.items()}

    # V1A/B/C: Tid vs poang per karttyp
    v1_paths: List[Path] = []
    for mode_name, color, tag in [
        ("Moving", "#2A77D4", "V1A"),
        ("No move", "#279B70", "V1B"),
        ("NMPZ", "#7A67D8", "V1C"),
    ]:
        part = dfo[dfo["mode3"] == mode_name].copy()
        fig, ax = plt.subplots(figsize=(8.6, 4.8))
        if not part.empty:
            ax.scatter(part["total_time"] / 60.0, part["total_pts"], s=14, alpha=0.14, color="#808B96", label="Alla rundor")
            by_player = (
                part.groupby("player", as_index=False)
                .agg(
                    mean_time_min=("total_time", lambda s: float(np.mean(s)) / 60.0),
                    mean_pts=("total_pts", "mean"),
                    maps=("total_pts", "count"),
                )
            )

            sizes = [22 + min(60.0, float(m) * 1.2) for m in by_player["maps"].tolist()]
            ax.scatter(
                by_player["mean_time_min"].tolist(),
                by_player["mean_pts"].tolist(),
                s=sizes,
                alpha=0.82,
                color=color,
                edgecolors="white",
                linewidths=0.5,
                label="Spelarmedel",
            )
            _annotate_all_points(
                ax,
                [float(x) for x in by_player["mean_time_min"].tolist()],
                [float(x) for x in by_player["mean_pts"].tolist()],
                [str(x) for x in by_player["player"].tolist()],
                fontsize=6,
            )
            ax.set_xlabel("Tid per karta (min)")
            ax.set_ylabel("GeoGuessr-poäng")
            ax.legend(loc="best", fontsize=7, frameon=True)
        else:
            _empty_plot(ax)
        ax.set_title(f"{tag}: Tid vs poäng ({mode_name})")
        v1_paths.append(_save_fig(fig, f"{tag}_tid_vs_poang_{mode_name.lower().replace(' ', '_')}.png"))

    # V2: Total rapoang for spelare med "mer an en full vecka"
    if expanded_players:
        top_raw = (
            df_total[df_total["player"].astype(str).isin(expanded_players)]
            .sort_values(["total_pts", "total_borda"], ascending=[False, False])
            .reset_index(drop=True)
        )
    else:
        top_raw = df_total.sort_values(["total_pts", "total_borda"], ascending=[False, False]).reset_index(drop=True)
    v2_labels = [str(x) for x in top_raw["player"].tolist()]
    v2_values = [float(x) for x in pd.to_numeric(top_raw["total_pts"], errors="coerce").fillna(0).tolist()]
    fig_h = max(6.0, min(13.0, 1.2 + 0.24 * max(1, len(v2_labels))))
    fig, ax = plt.subplots(figsize=(11.2, fig_h))
    if v2_values:
        labels_rev = list(reversed(v2_labels))
        vals_rev = list(reversed(v2_values))
        ys = list(range(len(labels_rev)))
        ax.barh(ys, vals_rev, color="#279B70")
        ax.set_yticks(ys)
        ax.set_yticklabels(labels_rev, fontsize=8)
        ax.set_ylabel("Spelare")
        ax.set_xlabel("GeoGuessr-poäng")
    else:
        _empty_plot(ax)
    ax.set_title("V2: Total råpoäng (spelare med mer än en full vecka)")
    v2_path = _save_fig(fig, "V2_total_rapoang_expanded.png")

    # V3: Snitt GeoGuessr-poang per karttyp och spelare (heatmap)
    mode_avg = (
        dfo[dfo["mode3"].isin(["Moving", "No move", "NMPZ"])]
        .groupby(["player", "mode3"], as_index=False)
        .agg(avg_pts=("total_pts", "mean"))
    )
    if expanded_players:
        v3_players = sorted(expanded_players, key=lambda p: total_pts_dict.get(p, 0.0), reverse=True)
    elif qualified_players:
        v3_players = sorted(qualified_players, key=lambda p: total_pts_dict.get(p, 0.0), reverse=True)
    else:
        v3_players = (
            dfo.groupby("player", as_index=False)
            .agg(maps=("week_map_key", "nunique"))
            .sort_values("maps", ascending=False)["player"]
            .astype(str)
            .head(40)
            .tolist()
        )
    v3_pivot = (
        mode_avg.pivot_table(index="player", columns="mode3", values="avg_pts", aggfunc="mean")
        .reindex(v3_players)
        .fillna(0.0)
    )
    for col in ["Moving", "No move", "NMPZ"]:
        if col not in v3_pivot.columns:
            v3_pivot[col] = 0.0
    v3_pivot = v3_pivot[["Moving", "No move", "NMPZ"]]
    fig_h = max(6.0, min(13.0, 1.8 + 0.20 * max(1, len(v3_pivot.index))))
    fig, ax = plt.subplots(figsize=(10.8, fig_h))
    if not v3_pivot.empty:
        im = ax.imshow(v3_pivot.values, aspect="auto", cmap="YlGnBu")
        ax.set_xticks([0, 1, 2])
        ax.set_xticklabels(["Moving", "No move", "NMPZ"])
        ax.set_yticks(list(range(len(v3_pivot.index))))
        ax.set_yticklabels([str(p) for p in v3_pivot.index], fontsize=7)
        fig.colorbar(im, ax=ax, fraction=0.035, pad=0.02, label="Snittpoäng")
    else:
        _empty_plot(ax)
    ax.set_title("V3: Snitt GeoGuessr-poäng per karttyp och spelare")
    v3_path = _save_fig(fig, "V3_snitt_ggpoang_karttyp_spelare.png")

    # V4: Aktiva spelare per vecka
    per_week_players = (
        dfo.groupby("week", as_index=False)
        .agg(active_players=("player", "nunique"))
        .set_index("week")
        .reindex(weeks_order, fill_value=0.0)
        .reset_index()
    )
    v4_labels = [str(x) for x in per_week_players["week"].tolist()]
    v4_values = [float(x) for x in pd.to_numeric(per_week_players["active_players"], errors="coerce").fillna(0).tolist()]
    fig, ax = plt.subplots(figsize=(8.6, 4.8))
    if v4_values:
        xs = list(range(len(v4_values)))
        ax.plot(xs, v4_values, marker="o", color="#E0862B")
        ax.set_xticks(xs)
        ax.set_xticklabels(v4_labels, rotation=30, ha="right")
        ax.set_ylabel("Antal spelare")
    else:
        _empty_plot(ax)
    ax.set_title("V4: Aktiva spelare per vecka")
    v4_path = _save_fig(fig, "V4_aktiva_spelare_vecka.png")

    cat_slots = {
        "Moving": ["moving_1", "moving_2"],
        "No move": ["no_move_1", "no_move_2"],
        "NMPZ": ["nmpz_1", "nmpz_2"],
        "Sverige": ["moving_1", "no_move_2"],
    }

    # V5: Ligans snitt GeoGuessr-poang per underliga-kategori
    v5_labels = ["Moving", "No move", "NMPZ", "Sverige"]
    v5_values: List[float] = []
    for cat in v5_labels:
        part = dfo[dfo["slot_key"].isin(cat_slots[cat])]
        v5_values.append(float(part["total_pts"].mean()) if not part.empty else 0.0)
    fig, ax = plt.subplots(figsize=(8.6, 4.8))
    if any(v5_values):
        xs = list(range(len(v5_values)))
        ax.bar(xs, v5_values, color=["#2A77D4", "#279B70", "#7A67D8", "#E0862B"])
        ax.set_xticks(xs)
        ax.set_xticklabels(v5_labels)
        ax.set_ylabel("Snitt GeoGuessr-poäng")
    else:
        _empty_plot(ax)
    ax.set_title("V5: Ligans snitt GeoGuessr-poäng per underliga-kategori")
    v5_path = _save_fig(fig, "V5_ligans_snitt_ggpoang_underliga.png")

    # V6: Topp-spelare per karttyp (snitt)
    by_mode_player = (
        dfo[dfo["mode3"].isin(["Moving", "No move", "NMPZ"])]
        .groupby(["player", "mode3"], as_index=False)
        .agg(avg_pts=("total_pts", "mean"))
    )
    v6_players = (
        df_total.sort_values(["total_pts", "total_borda"], ascending=[False, False])["player"]
        .astype(str).head(12).tolist()
    )
    v6_pivot = (
        by_mode_player.pivot_table(index="player", columns="mode3", values="avg_pts", aggfunc="mean")
        .reindex(v6_players)
        .fillna(0.0)
    )
    for col in ["Moving", "No move", "NMPZ"]:
        if col not in v6_pivot.columns:
            v6_pivot[col] = 0.0
    v6_pivot = v6_pivot[["Moving", "No move", "NMPZ"]]
    fig, ax = plt.subplots(figsize=(8.6, 4.8))
    if not v6_pivot.empty:
        xs = np.arange(len(v6_pivot.index))
        w = 0.26
        ax.bar(xs - w, v6_pivot["Moving"].tolist(), width=w, label="Moving", color="#2A77D4")
        ax.bar(xs, v6_pivot["No move"].tolist(), width=w, label="No move", color="#279B70")
        ax.bar(xs + w, v6_pivot["NMPZ"].tolist(), width=w, label="NMPZ", color="#7A67D8")
        ax.set_xticks(xs)
        ax.set_xticklabels([str(x) for x in v6_pivot.index], rotation=30, ha="right", fontsize=8)
        ax.set_ylabel("Snitt GeoGuessr-poäng")
        ax.legend(fontsize=8)
    else:
        _empty_plot(ax)
    ax.set_title("V6: Topp-spelare per karttyp (snitt GeoGuessr-poäng)")
    v6_path = _save_fig(fig, "V6_toppspelare_karttyp.png")

    # V7: Poangfordelning per karttyp (boxplot)
    fig, ax = plt.subplots(figsize=(8.6, 4.8))
    box_data = [
        dfo[dfo["mode3"] == "Moving"]["total_pts"].tolist(),
        dfo[dfo["mode3"] == "No move"]["total_pts"].tolist(),
        dfo[dfo["mode3"] == "NMPZ"]["total_pts"].tolist(),
    ]
    if any(len(b) > 0 for b in box_data):
        ax.boxplot(box_data, labels=["Moving", "No move", "NMPZ"], showfliers=False)
        ax.set_ylabel("GeoGuessr-poäng")
    else:
        _empty_plot(ax)
    ax.set_title("V7: Poängfördelning per karttyp")
    v7_path = _save_fig(fig, "V7_poangfordelning_karttyp_boxplot.png")

    # V8: Stabilitet vs niva (std mot snitt)
    stab = (
        dfo.groupby("player", as_index=False)
        .agg(
            mean_pts=("total_pts", "mean"),
            std_pts=("total_pts", "std"),
            maps=("week_map_key", "nunique"),
        )
    )
    stab["std_pts"] = pd.to_numeric(stab["std_pts"], errors="coerce").fillna(0.0)
    fig, ax = plt.subplots(figsize=(11.0, 6.6))
    if not stab.empty:
        ax.scatter(stab["std_pts"].tolist(), stab["mean_pts"].tolist(), color="#2A77D4", alpha=0.6, s=30)
        _annotate_all_points(
            ax,
            [float(x) for x in stab["std_pts"].tolist()],
            [float(x) for x in stab["mean_pts"].tolist()],
            [str(x) for x in stab["player"].tolist()],
            fontsize=6,
        )
        ax.set_xlabel("Standardavvikelse i poäng")
        ax.set_ylabel("Snitt GeoGuessr-poäng")
    else:
        _empty_plot(ax)
    ax.set_title("V8: Stabilitet vs nivå (std mot snittpoäng)")
    v8_path = _save_fig(fig, "V8_stabilitet_vs_niva.png")

    # V9: Veckotrend-heatmap
    week_player = (
        dfo.groupby(["player", "week"], as_index=False)
        .agg(avg_pts=("total_pts", "mean"))
    )
    if expanded_players:
        v9_players = sorted(expanded_players, key=lambda p: total_pts_dict.get(p, 0.0), reverse=True)
    elif qualified_players:
        v9_players = sorted(qualified_players, key=lambda p: total_pts_dict.get(p, 0.0), reverse=True)
    else:
        v9_players = (
            dfo.groupby("player", as_index=False)
            .agg(maps=("week_map_key", "nunique"))
            .sort_values("maps", ascending=False)["player"]
            .astype(str).head(40).tolist()
        )
    v9_pivot = (
        week_player.pivot_table(index="player", columns="week", values="avg_pts", aggfunc="mean")
        .reindex(index=v9_players, columns=weeks_order)
    )
    fig_h = max(6.0, min(14.0, 1.8 + 0.18 * max(1, len(v9_players))))
    fig_w = max(10.6, 7.4 + 0.9 * max(1, len(weeks_order)))
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    if not v9_pivot.empty:
        centered = v9_pivot.sub(v9_pivot.mean(axis=1), axis=0).fillna(0.0)
        im = ax.imshow(centered.values, aspect="auto", cmap="RdYlGn")
        ax.set_xticks(list(range(len(weeks_order))))
        ax.set_xticklabels([str(w) for w in weeks_order], rotation=30, ha="right", fontsize=8)
        ax.set_yticks(list(range(len(centered.index))))
        ax.set_yticklabels([str(p) for p in centered.index], fontsize=7)
        fig.colorbar(im, ax=ax, fraction=0.035, pad=0.02, label="Över/under eget snitt")
    else:
        _empty_plot(ax)
    ax.set_title("V9: Veckotrend-heatmap (över/under eget snitt)")
    v9_path = _save_fig(fig, "V9_veckotrend_heatmap.png")

    # V10: PCA med loadings
    feat_rows: List[dict] = []
    for player, grp in dfo.groupby("player"):
        entry = {"player": str(player)}
        for cat, slots in cat_slots.items():
            sub = grp[grp["slot_key"].isin(slots)]
            entry[f"avg_pts_{cat}"] = float(sub["total_pts"].mean()) if not sub.empty else 0.0
        entry["avg_time_min"] = float(grp["total_time"].mean()) / 60.0 if not grp.empty else 0.0
        entry["std_pts"] = float(grp["total_pts"].std()) if len(grp) > 1 else 0.0
        entry["maps"] = int(grp["week_map_key"].nunique())
        feat_rows.append(entry)
    feat_df = pd.DataFrame(feat_rows)
    feature_cols = ["avg_pts_Moving", "avg_pts_No move", "avg_pts_NMPZ", "avg_pts_Sverige", "avg_time_min", "std_pts"]

    fig = plt.figure(figsize=(12.6, 7.1))
    gs = fig.add_gridspec(2, 2, width_ratios=[2.1, 1.0], wspace=0.34, hspace=0.34)
    ax_sc = fig.add_subplot(gs[:, 0])
    ax_l1 = fig.add_subplot(gs[0, 1])
    ax_l2 = fig.add_subplot(gs[1, 1])

    if not feat_df.empty and len(feat_df) >= 3:
        if expanded_players:
            feat_df = feat_df[feat_df["player"].astype(str).isin(expanded_players)].copy()
        elif qualified_players:
            feat_df = feat_df[feat_df["player"].astype(str).isin(qualified_players)].copy()
        if feat_df.empty:
            feat_df = pd.DataFrame(feat_rows)
        x_mat = feat_df[feature_cols].to_numpy(dtype=float)
        x_center = x_mat - x_mat.mean(axis=0, keepdims=True)
        scale = x_center.std(axis=0, keepdims=True)
        scale[scale == 0] = 1.0
        x_norm = x_center / scale

        _, svals, vt = np.linalg.svd(x_norm, full_matrices=False)
        pcs = x_norm @ vt.T[:, :2]
        feat_df["pc1"] = pcs[:, 0]
        feat_df["pc2"] = pcs[:, 1]

        exp = svals * svals
        exp_ratio = exp / exp.sum() if exp.sum() > 0 else exp
        pc1_pct = float(exp_ratio[0] * 100.0) if len(exp_ratio) > 0 else 0.0
        pc2_pct = float(exp_ratio[1] * 100.0) if len(exp_ratio) > 1 else 0.0

        colors = [float(total_pts_dict.get(str(p), 0.0)) for p in feat_df["player"].tolist()]
        sc = ax_sc.scatter(feat_df["pc1"], feat_df["pc2"], c=colors, cmap="viridis", s=50, alpha=0.85)
        fig.colorbar(sc, ax=ax_sc, fraction=0.035, pad=0.02, label="Total råpoäng")

        _annotate_all_points(
            ax_sc,
            [float(x) for x in feat_df["pc1"].tolist()],
            [float(x) for x in feat_df["pc2"].tolist()],
            [str(x) for x in feat_df["player"].tolist()],
            fontsize=6,
        )
        ax_sc.set_xlabel(f"PC1 ({pc1_pct:.1f}% förklarad varians)")
        ax_sc.set_ylabel(f"PC2 ({pc2_pct:.1f}% förklarad varians)")
        ax_sc.set_title("Spelare i PCA-rum")

        load_pc1 = list(zip(feature_cols, vt[0, :len(feature_cols)]))
        load_pc2 = list(zip(feature_cols, vt[1, :len(feature_cols)] if vt.shape[0] > 1 else np.zeros(len(feature_cols))))
        load_pc1 = sorted(load_pc1, key=lambda x: abs(float(x[1])), reverse=True)
        load_pc2 = sorted(load_pc2, key=lambda x: abs(float(x[1])), reverse=True)

        y1 = list(range(len(load_pc1)))
        ax_l1.barh(y1, [float(x[1]) for x in load_pc1], color="#2A77D4")
        ax_l1.set_yticks(y1)
        ax_l1.set_yticklabels([str(x[0]) for x in load_pc1], fontsize=7)
        ax_l1.invert_yaxis()
        ax_l1.set_title("PC1 bidrag (loading)")

        y2 = list(range(len(load_pc2)))
        ax_l2.barh(y2, [float(x[1]) for x in load_pc2], color="#279B70")
        ax_l2.set_yticks(y2)
        ax_l2.set_yticklabels([str(x[0]) for x in load_pc2], fontsize=7)
        ax_l2.invert_yaxis()
        ax_l2.set_title("PC2 bidrag (loading)")
    else:
        _empty_plot(ax_sc)
        _empty_plot(ax_l1)
        _empty_plot(ax_l2)
    fig.suptitle("V10: PCA spelarmönster + bidrag till PC1/PC2", fontsize=12, y=0.99)
    v10_path = _save_fig(fig, "V10_pca_spelarmonster_loadings.png")

    # V11: Standardavvikelse per karttyp och spelare
    std_by_mode = (
        dfo[dfo["mode3"].isin(["Moving", "No move", "NMPZ"])]
        .groupby(["player", "mode3"], as_index=False)
        .agg(std_pts=("total_pts", "std"))
    )
    std_by_mode["std_pts"] = pd.to_numeric(std_by_mode["std_pts"], errors="coerce").fillna(0.0)
    if expanded_players:
        v11_players = sorted(expanded_players, key=lambda p: total_pts_dict.get(p, 0.0), reverse=True)
    elif qualified_players:
        v11_players = sorted(qualified_players, key=lambda p: total_pts_dict.get(p, 0.0), reverse=True)
    else:
        v11_players = (
            dfo.groupby("player", as_index=False)
            .agg(maps=("week_map_key", "nunique"))
            .sort_values("maps", ascending=False)["player"]
            .astype(str).head(40).tolist()
        )
    v11_pivot = (
        std_by_mode.pivot_table(index="player", columns="mode3", values="std_pts", aggfunc="mean")
        .reindex(v11_players)
        .fillna(0.0)
    )
    for col in ["Moving", "No move", "NMPZ"]:
        if col not in v11_pivot.columns:
            v11_pivot[col] = 0.0
    v11_pivot = v11_pivot[["Moving", "No move", "NMPZ"]]
    fig_h = max(6.0, min(13.0, 1.8 + 0.19 * max(1, len(v11_players))))
    fig, ax = plt.subplots(figsize=(10.8, fig_h))
    if not v11_pivot.empty:
        im = ax.imshow(v11_pivot.values, aspect="auto", cmap="OrRd")
        ax.set_xticks([0, 1, 2])
        ax.set_xticklabels(["Moving", "No move", "NMPZ"])
        ax.set_yticks(list(range(len(v11_pivot.index))))
        ax.set_yticklabels([str(p) for p in v11_pivot.index], fontsize=7)
        fig.colorbar(im, ax=ax, fraction=0.035, pad=0.02, label="Std i poäng")
    else:
        _empty_plot(ax)
    ax.set_title("V11: Standardavvikelse per karttyp och spelare")
    v11_path = _save_fig(fig, "V11_std_per_karttyp_spelare.png")

    # Place images lower and larger so overview text remains visible and plots are easier to read.
    anchors: List[str] = []
    first_row = 24
    row_step = 31
    for i in range(7):
        r = first_row + i * row_step
        anchors.append(f"A{r}")
        anchors.append(f"N{r}")

    image_specs: List[Tuple[Path, int, int]] = [
        (v1_paths[0], 760, 430),
        (v1_paths[1], 760, 430),
        (v1_paths[2], 760, 430),
        (v2_path, 760, 520),
        (v3_path, 760, 520),
        (v4_path, 760, 430),
        (v5_path, 760, 430),
        (v6_path, 760, 430),
        (v7_path, 760, 430),
        (v8_path, 760, 430),
        (v9_path, 760, 520),
        (v10_path, 760, 500),
        (v11_path, 760, 520),
    ]

    for idx, (img_path, w, h) in enumerate(image_specs):
        if idx >= len(anchors):
            break
        _insert_image(img_path, anchors[idx], width=w, height=h)

    info_row = first_row + 7 * row_step + 3
    ws[f"A{info_row}"] = f"Bilder sparade i: {image_dir}"
    ws[f"A{info_row}"].font = Font(color="4F4F4F", italic=True)
    ws[f"A{info_row+1}"] = (
        "Namn i scatter: alla spelare visas. "
        f"Utökad urvalsregel i V2/V9/V11: full vecka + extra spel (>=7 kartor över minst 2 veckor)."
    )
    ws[f"A{info_row+1}"].font = Font(color="4F4F4F", italic=True)


def write_underligor_sheet(wb: Workbook, df_overview: pd.DataFrame) -> None:
    ws = wb.create_sheet("Underligor")
    leagues = ["Moving", "No move", "NMPZ", "Sverige"]
    block_widths = [4.5, 22.0, 12.0, 10.0, 8.0, 8.0]
    block_cols = len(block_widths)
    gap_cols = 1
    total_cols = (block_cols + gap_cols) * len(leagues) - gap_cols
    merge_and_style(ws, 1, 1, 1, total_cols, "Underligor", fill=DARK, font=FONT_HDR_BIG, align=CENTER)

    widths: Dict[int, float] = {}
    for i in range(len(leagues)):
        start_col = 1 + i * (block_cols + gap_cols)
        for offset, w in enumerate(block_widths):
            widths[start_col + offset] = w
        if i < len(leagues) - 1:
            widths[start_col + block_cols] = 3.0
    set_col_widths(ws, widths)

    tables = compute_subleague_tables(df_overview)
    section_row = 3
    headers = ["#", "Spelare", "Poäng", "Total pts", "Kartor", "Veckor"]

    for i, league_name in enumerate(leagues):
        start_col = 1 + i * (block_cols + gap_cols)
        end_col = start_col + block_cols - 1
        merge_and_style(ws, section_row, start_col, section_row, end_col, league_name, fill=MID, font=FONT_HDR_MED, align=CENTER)

        header_row = section_row + 1
        for j, h in enumerate(headers):
            c = start_col + j
            ws.cell(header_row, c).value = h
            style_cell(ws, header_row, c, fill=MID, font=FONT_HDR, align=CENTER)

        data_start_row = section_row + 2
        table = tables.get(league_name, pd.DataFrame())
        for idx, row in enumerate(table.itertuples(index=False), start=1):
            r = data_start_row + (idx - 1)
            base_fill = ROW_A if (idx % 2 == 1) else ROW_B
            fill = rank_row_fill(idx, base_fill)
            ws.cell(r, start_col + 0).value = idx
            ws.cell(r, start_col + 1).value = row.player
            ws.cell(r, start_col + 2).value = float(row.league_points)
            ws.cell(r, start_col + 3).value = int(row.total_pts)
            ws.cell(r, start_col + 4).value = int(row.maps_counted)
            ws.cell(r, start_col + 5).value = int(row.weeks_counted)

            for c in range(start_col, end_col + 1):
                align = LEFT if c == 2 else CENTER
                if c == start_col + 1:
                    align = LEFT
                font = Font(color="000000", bold=True) if c == start_col + 2 else FONT_BODY
                style_cell(ws, r, c, fill=fill, font=font, align=align)

    ws.freeze_panes = "A5"


def write_information_sheet(wb: Workbook, info_rows: Optional[List[str]] = None) -> None:
    ws = wb.create_sheet("Information")

    merge_and_style(ws, 1, 1, 2, 2, "Information", fill=DARK, font=FONT_HDR_BIG, align=CENTER)

    rows = _normalize_information_rows(info_rows if info_rows is not None else default_information_rows())

    set_col_widths(ws, {1: 4.5, 2: 185.0})
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 40

    for i, text in enumerate(rows, start=0):
        r = 3 + i
        fill = ROW_A if (i % 2 == 0) else ROW_B
        ws.cell(r, 1).value = "•"
        ws.cell(r, 2).value = text
        style_cell(ws, r, 1, fill=fill, font=FONT_BODY, align=CENTER)
        style_cell(ws, r, 2, fill=fill, font=FONT_BODY, align=LEFT)
        ws.row_dimensions[r].height = 34


def write_raw_sheet(wb: Workbook, df_overview: pd.DataFrame) -> None:
    ws = wb.create_sheet("Raw")
    if df_overview.empty:
        ws["A1"].value = "No data"
        return

    headers = list(df_overview.columns)
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c).value = h
        style_cell(ws, 1, c, fill=DARK, font=FONT_HDR, align=CENTER)

    for r_idx, (_, row) in enumerate(df_overview.iterrows(), start=2):
        for c_idx, h in enumerate(headers, start=1):
            v = row[h]
            if pd.isna(v):
                v = ""
            ws.cell(r_idx, c_idx).value = v
            style_cell(ws, r_idx, c_idx, fill=WHITE, font=FONT_BODY, align=LEFT if h == "player" else CENTER)

    ws.freeze_panes = "A2"
    for c_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(len(str(h)) + 2, 10), 40)


def _output_with_suffix(path: Path, index: int) -> Path:
    if index <= 0:
        return path
    return path.with_name(f"{path.stem} ({index}){path.suffix}")


def save_workbook_with_fallback(wb: Workbook, desired_path: Path, max_attempts: int = 50) -> Path:
    """
    Save workbook to desired_path, overwriting when possible.
    If the target is locked (e.g. opened in Excel), fallback to suffixed names.
    """
    last_exc: Optional[Exception] = None
    for n in range(0, max_attempts + 1):
        candidate = _output_with_suffix(desired_path, n)
        try:
            wb.save(candidate)
            if n > 0:
                print(f"[WARN] output file locked: {desired_path.name}. Saved as {candidate.name}")
            return candidate
        except PermissionError as e:
            last_exc = e
            continue
    if last_exc is not None:
        raise last_exc
    raise RuntimeError(f"Could not save workbook: {desired_path}")


# ============================================================
# Main
# ============================================================

def parse_week_specs(week_args: List[str]) -> List[WeekSpec]:
    if not week_args:
        raise SystemExit(
            'No weeks specified. Example:\n'
            '  python geoguessr_league_build_xlsx.py --week "Vecka 1|urls_week1.txt|2026-02-18 20:00" --week "Vecka 2|urls_week2.txt|2026-02-25 20:00"'
        )

    out: List[WeekSpec] = []
    for s in week_args:
        parts = [p.strip() for p in s.split("|")]
        if len(parts) < 2:
            raise SystemExit(f'Bad --week "{s}". Expected "LABEL|URLS_FILE|DEADLINE(optional)".')
        label = parts[0]
        urls_path = Path(parts[1]).expanduser()
        deadline = parts[2] if len(parts) >= 3 and parts[2] else None
        out.append(WeekSpec(label=label, urls_path=urls_path, deadline=deadline))
    return out


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    weeks = parse_week_specs(args.week)

    ncfa = (args.ncfa or os.environ.get("GEOGUESSR_NCFA", "")).strip()
    if not ncfa:
        raise SystemExit("Missing _ncfa. Set GEOGUESSR_NCFA or pass --ncfa.")

    print("[START] python:", sys.executable)
    print("[START] cwd   :", Path.cwd())
    print("[START] weeks :", [(w.label, str(w.urls_path), w.deadline) for w in weeks])
    print("[START] fetch_played_at:", bool(args.fetch_played_at))
    print("[START] tz:", args.tz)

    if args.information_config.strip():
        info_config_path = Path(args.information_config.strip()).expanduser()
    else:
        info_config_path = Path(DEFAULT_INFORMATION_CONFIG_NAME)
    info_rows = load_information_rows(info_config_path, debug=args.debug)
    print("[START] information_config:", info_config_path if info_config_path.exists() else "(default built-in)")

    session = make_session(ncfa)

    # collect deadlines (epoch)
    deadlines_epoch: Dict[str, int] = {}
    for w in weeks:
        if w.deadline:
            deadlines_epoch[w.label] = parse_deadline_epoch(w.deadline, args.tz)

    # Build all entries
    all_entries: List[Entry] = []
    all_map_meta: List[dict] = []
    any_played_at = False
    successful_weeks: List[WeekSpec] = []
    failed_weeks: List[Tuple[str, str]] = []

    for w in weeks:
        try:
            entries, week_map_meta, has_any, failed_maps = build_week_entries(
                session=session,
                week=w,
                tz_name=args.tz,
                timeout=args.timeout,
                debug=args.debug,
                dump_json=args.dump_json,
                page_size=args.page_size,
                max_players=args.max_players,
                fetch_played_at=bool(args.fetch_played_at and (w.deadline is not None)),
            )
            all_entries.extend(entries)
            all_map_meta.extend(week_map_meta)
            any_played_at = any_played_at or has_any
            successful_weeks.append(w)
            print(f"[OK] built entries for {w.label}: {len(entries)} rows")
            if failed_maps > 0:
                print(f"[WARN] {w.label}: {failed_maps} map(ar) kunde inte hämtas och hoppades över.")
        except Exception as e:
            failed_weeks.append((w.label, str(e)))
            print(f"[WARN] hoppar över vecka {w.label}: {e}")
            continue

    if not successful_weeks:
        raise SystemExit("Kunde inte bearbeta någon vecka. Kontrollera URL-filer, _ncfa och nätverksåtkomst.")

    if failed_weeks:
        print("[WARN] följande veckor kunde inte bearbetas:")
        for label, err in failed_weeks:
            print(f"  - {label}: {err}")

    # Compute tables for ALL (unfiltered)
    df_overview_all, df_weekly_all, df_meta_all = compute_week_tables(all_entries, tie_mode=args.tie, map_meta_rows=all_map_meta)
    df_total_all, df_stats_all = compute_total_tables(df_overview_all)

    # Decide filtering
    can_filter = bool(deadlines_epoch) and bool(args.fetch_played_at) and any_played_at

    # Build filtered data (if possible)
    df_overview_f = df_weekly_f = df_meta_f = df_total_f = df_stats_f = None
    if can_filter:
        now_epoch = int(time.time())
        open_weeks = [w for w, dl in deadlines_epoch.items() if dl > now_epoch]
        if open_weeks:
            print(f"[FILTER] open week(s), deadline not reached yet: {open_weeks}. Those weeks are not filtered.")

        filtered_entries = filter_entries_by_deadlines(
            all_entries,
            deadlines_epoch,
            keep_missing_time=bool(args.keep_missing_time),
            now_epoch=now_epoch,
        )
        df_overview_f, df_weekly_f, df_meta_f = compute_week_tables(filtered_entries, tie_mode=args.tie, map_meta_rows=all_map_meta)
        df_total_f, df_stats_f = compute_total_tables(df_overview_f)
        print(f"[FILTER] enabled. Filtered rows: {len(filtered_entries)} (from {len(all_entries)})")
    else:
        if deadlines_epoch and args.fetch_played_at:
            print("[FILTER] deadlines provided, but could not extract any played_at timestamps from API. Will write only ALL file.")
        elif deadlines_epoch and not args.fetch_played_at:
            print("[FILTER] deadlines provided, but --fetch-played-at not enabled. Will write only ALL file.")
        else:
            print("[FILTER] no deadlines -> will write only ALL file.")

    # Write ALL workbook
    out_all = Path(f"{args.out_base}_all.xlsx")
    wb_all = Workbook()
    # remove default sheet
    wb_all.remove(wb_all.active)
    write_information_sheet(wb_all, info_rows)

    week_labels = [w.label for w in successful_weeks]

    # Week tabs (ALL)
    for w in successful_weeks:
        dl_str = w.deadline or ""
        write_week_sheet(wb_all, w.label, f"Deadline {dl_str}" if dl_str else "Deadline", df_weekly_all, df_overview_all, df_meta_all)

    write_total_sheet(wb_all, df_total_all, df_overview_all, week_labels)
    write_stats_sheet(wb_all, df_stats_all)
    write_underligor_sheet(wb_all, df_overview_all)
    write_visualizations_sheet(
        wb_all,
        df_overview_all,
        df_total_all,
        week_labels,
        image_dir=out_all.parent / "visualizations" / out_all.stem,
    )
    write_raw_sheet(wb_all, df_overview_all)

    actual_out_all = save_workbook_with_fallback(wb_all, out_all)
    print("[DONE] wrote:", actual_out_all)

    # Write FILTERED workbook (if available)
    if can_filter and df_overview_f is not None and df_weekly_f is not None and df_meta_f is not None and df_total_f is not None and df_stats_f is not None:
        out_f = Path(f"{args.out_base}_filtered.xlsx")
        wb_f = Workbook()
        wb_f.remove(wb_f.active)
        write_information_sheet(wb_f, info_rows)

        for w in successful_weeks:
            dl_str = w.deadline or ""
            write_week_sheet(wb_f, w.label, f"Deadline {dl_str}" if dl_str else "Deadline", df_weekly_f, df_overview_f, df_meta_f)

        write_total_sheet(wb_f, df_total_f, df_overview_f, week_labels)
        write_stats_sheet(wb_f, df_stats_f)
        write_underligor_sheet(wb_f, df_overview_f)
        write_visualizations_sheet(
            wb_f,
            df_overview_f,
            df_total_f,
            week_labels,
            image_dir=out_f.parent / "visualizations" / out_f.stem,
        )
        write_raw_sheet(wb_f, df_overview_f)

        actual_out_f = save_workbook_with_fallback(wb_f, out_f)
        print("[DONE] wrote:", actual_out_f)

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except SystemExit:
        raise
    except Exception:
        print("[FATAL] Unhandled exception")
        traceback.print_exc()
        raise
